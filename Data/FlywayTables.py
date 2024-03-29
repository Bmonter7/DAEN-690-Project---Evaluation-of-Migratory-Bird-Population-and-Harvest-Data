from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.styles import NamedStyle
import pandas as pd
from pandas import DataFrame
import sys


def create_table_to_ws(workbook: Workbook, 
                 estimate_data_df: DataFrame, average_data_df: DataFrame, 
                 asterisk_text_list,
                 table_title: str,
                 sheet_name: str): 
    
    print_info('Begin creating table for table [' + str(table_title) + '] on sheet ' + sheet_name)
    ws1 = workbook.create_sheet(sheet_name)
    
    #track starting row for the table in sheet
    table_start = 0

    #track the starting row for the table average section in sheet
    table_avg_start = 0

    #track the table height
    table_height = 0

    #track the table header
    table_headers = list(estimate_data_df.columns)

    #track the table width
    estimate_data_df[' '] = ''
    average_data_df[' '] = ''

    table_width = len(list(estimate_data_df.columns))
    estimate_data_df.iloc[:, 0] = estimate_data_df.iloc[:, 0].astype(str)
    average_data_df.iloc[:, 0] = average_data_df.iloc[:, 0].astype(str)

    #add column headings. NB. these must be strings
    ws1.append([table_title])
    table_start += 1
    set_title_row_style(ws1[table_start])

    #merge title row
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=table_width)
    ws1.append(table_headers)
    table_height += 1
    table_start += 1

    #setting font and alignment for table header
    for idx, cell in enumerate(ws1[table_start]):
        if (idx > 0):
            cell.alignment = Alignment(horizontal='right')
        cell.font = Font(bold=True)

    #adding estimate data to ws
    for row in estimate_data_df.to_numpy():
        ws1.append(row.tolist())
        set_data_row_style(ws1[table_start+table_height])
        table_height += 1
        


    
    
    #Add 'Average' section title
    ws1.append(['Averages'])
    table_height += 1
    table_avg_start = table_start+table_height

    #Add 'Average' data section
    for row in average_data_df.to_numpy():
        ws1.append(row.tolist())
        set_data_row_style(ws1[table_start+table_height])
        table_height += 1
        
    

    #set border for the entire table
    set_border_rows(ws1[table_start:table_start+table_height-1])

    #set border for header section
    set_border_rows(ws1[table_start], True)

    #set border for average section
    set_border_rows(ws1[table_avg_start-1:table_start+table_height-1]) 

    print_info('Table Created with Size='+str(table_width)+'x' + str(table_height))

    asterisk_start_row = table_start+table_height
    #setting the * section text
    for item in asterisk_text_list:
        ws1.append([item])
        ws1.merge_cells(start_row=asterisk_start_row, start_column=1, end_row=asterisk_start_row, end_column=table_width)
        ws1.cell(asterisk_start_row, 1).font = Font(size=10)
        asterisk_start_row += 1

    print_info('Done creating Excel table for [' + table_title + '] on sheet ' + sheet_name)


def _draw_border(row, pos_y, max_x, max_y):
        side = Side(border_style='medium', color="FF000000")
        for pos_x, cell in enumerate(row):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border

def set_border_rows(rows, isOneRow=False):
    
    if (isOneRow):
        _draw_border(rows, 0, len(rows)-1, 0)
        return


    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    
    for pos_y, row in enumerate(rows):
        max_x = len(row) - 1  # index of the last cell
        _draw_border(row, pos_y, max_x, max_y)
        
    
#set style on a single row
def set_title_row_style(row):
    for cell in row:
        cell.font = Font(bold=True)

def set_data_row_style(row):
    for idx, cell in enumerate(row):
        try:
            if (idx > 0 and int(cell.value) > 0):
                cell.number_format = '#,###'
        except:
            continue
            


''' ########### FUNCTIONS: Printing Output ########### '''

def print_info(output_str):
    print("[INFO] " + str(output_str))

def print_error(output_str):
    print("[ERROR] " + str(output_str))

def print_fatal_exit(output_str):
    print("[FATAL] " + str(output_str))
    print("[FATAL] Exiting now...")
    sys.exit()



